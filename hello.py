#!/usr/bin/env python
"""
Read a mitmproxy dump file and process requests based on a tracker list or personal info list.
"""
import excel_IO
import pprint
import sys
import json
from mitmproxy import http
from mitmproxy import io
from mitmproxy.exceptions import FlowReadException
from urllib.parse import parse_qs
import os
import re
import openpyxl

os.system("")  # Enable ANSI escape sequences in Windows PowerShell

def validate_and_read_file():
    if len(sys.argv) != 2:
        print("Error: dump file required!")
        print(f"Usage: {sys.argv[0]} mitm_dump_file")
        exit(1)
    return sys.argv[1]

def process_request(f, trackerList=None, prsnlList=None):
    if not isinstance(f, http.HTTPFlow):
        return
    
    request = f.request
    method = request.method
    host = request.host

    if trackerList and match_tracker_list(trackerList, host):
        print_request_details(request, method, host)
        if method == "POST":
            process_post_request(request)
        print("\n!===============================!\n")

    if prsnlList:
        match_found = check_personal_info(request, prsnlList)
        if match_found:
            print_request_details(request, method, host)
            if method == "POST":
                process_post_request(request)
            print("\n!===============================!\n")
            write_to_excel(host, prepare_data_for_excel(request, method, host))

def print_request_details(request, method, host):
    print("\n!===============================!\n")
    if method == 'GET':
        print("\033[95mGET\033[0m")
    print(f"{method} {host}")
    print(f"path: {request.path}")
    print_query_parameters(request.query.items())
    print_headers(request.headers.items())
    print("\n----------------------\n")

def print_query_parameters(queries):
    if queries:
        print("queries: [")
        for k, v in queries:
            print(f"\t{k}={v}")
        print("]")
    else:
        print("queries: []")

def print_headers(headers):
    if headers:
        print("headers: [")
        for k, v in headers:
            print(f"\t{k}: {v}")
        print("]")
    else:
        print("headers: []")

def process_post_request(request):
    content_type, content_length = get_content_type_and_length(request.headers.items())
    if content_length > 0:
        print(f"=> {content_type}: [")
        process_post_data(request, content_type)
        print("]")
        print("\033[93m어떤 내용이 있다\033[0m")
    else:
        print("\033[93mPOST이긴 하다\033[0m")

def process_post_data(request, content_type):
    text = request.get_text(False)
    match content_type:
        case "application/x-www-form-urlencoded":
            data = parse_qs(text)
            for k, v in data.items():
                print(f"\t{k}={v}")
        case "application/json":
            data = json.loads(text)
            print(json.dumps(data, indent=4))
        case "text/plain":
            print(text)
        case _:
            print("\t** SKIP **")

def get_content_type_and_length(headers):
    content_type = ""
    content_length = 0
    for k, v in headers:
        if k.casefold() == "Content-Type".casefold():
            content_type = v.split(";")[0]
        elif k.casefold() == "Content-Length".casefold():
            content_length = int(v)
    return content_type, content_length

def match_tracker_list(trackerList, host):
    for pattern in trackerList:
        word_pattern = rf'{re.escape(pattern)}\b'
        if re.search(word_pattern, host, re.IGNORECASE):
            return True
    return False

def check_personal_info(request, prsnlList):
    if match_prsnl_list(prsnlList, request.query.items()):
        return True

    content_type, content_length = get_content_type_and_length(request.headers.items())
    if match_prsnl_list(prsnlList, request.headers.items()):
        return True

    if request.method == "POST" and content_length > 0:
        text = request.get_text(False)
        return match_post_data(content_type, prsnlList, text)

    return False

def match_prsnl_list(prsnlList, kv):
    if isinstance(kv, list):
        for k, v in kv:
            if match_pattern_in_prsnl_list(prsnlList, k, v):
                return True
    elif isinstance(kv, tuple) and len(kv) == 2:
        k, v = kv
        if match_pattern_in_prsnl_list(prsnlList, k, v):
            return True
    elif isinstance(kv, str):
        if match_pattern_in_prsnl_list(prsnlList, kv):
            return True
    return False

def match_pattern_in_prsnl_list(prsnlList, k, v=None):
    for pattern in prsnlList:
        if pattern is None:
            continue
        word_pattern = rf'{re.escape(str(pattern))}\b'
        if re.search(word_pattern, k, re.IGNORECASE) or (v and re.search(word_pattern, v, re.IGNORECASE)):
            return True
    return False

def match_post_data(content_type, prsnlList, text):
    match content_type:
        case "application/x-www-form-urlencoded":
            data = parse_qs(text)
            return match_prsnl_list(prsnlList, data.items())
        case "application/json":
            data = json.loads(text)
            return match_prsnl_list(prsnlList, json.dumps(data, indent=4))
        case "text/plain":
            return match_prsnl_list(prsnlList, text)
        case _:
            return False

def prepare_data_for_excel(request, method, host):
    data_to_write = [method, host, f"path: {request.path}"]
    queries = request.query.items()
    data_to_write.append("queries: [" + ", ".join([f"{k}={v}" for k, v in queries]) + "]" if queries else "queries: []")
    headers = request.headers.items()
    data_to_write.append("headers: [" + ", ".join([f"{k}: {v}" for k, v in headers]) + "]" if headers else "headers: []")
    return data_to_write

def process_flows(logfile_name, mode):
    try:
        with open(logfile_name, "rb") as logfile:
            freader = io.FlowReader(logfile)
            if mode == '1':
                trackerList = excel_IO.excel_trackerList_input()
                for f in freader.stream():
                    process_request(f, trackerList=trackerList)
            elif mode == '2':
                prsnlList = excel_IO.excel_prsnlList_input()
                for f in freader.stream():
                    process_request(f, prsnlList=prsnlList)
    except FlowReadException as e:
        print(f"Flow file corrupted: {e}")

def write_to_excel(host, data):
    excel_file_path = r"C:\Users\kfri1\Desktop\12output.xlsx"
    wb = open_or_create_workbook(excel_file_path)
    ws = wb[host] if host in wb.sheetnames else wb.create_sheet(title=host)
    last_column = ws.max_column + 1

    for i, value in enumerate(data, start=1):
        ws.cell(row=i, column=last_column, value=value)

    wb.save(excel_file_path)

def open_or_create_workbook(excel_file_path):
    return openpyxl.load_workbook(excel_file_path) if os.path.exists(excel_file_path) else openpyxl.Workbook()

def main():
    mode = input("숫자를 입력: ")
    logfile_name = validate_and_read_file()
    process_flows(logfile_name, mode)

if __name__ == "__main__":
    main()