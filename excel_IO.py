import openpyxl
import re
import os
from openpyxl.utils.exceptions import IllegalCharacterError

import gzip
import zlib
from io import BytesIO
import json

import ast


MAX_HOSTNAME_LENGTH = 31

def excel_trackerList_input():
    trackerList_path = r"C:\Users\xten\Desktop\goodchoicetracker.xlsx"
    wb = openpyxl.load_workbook(trackerList_path)
    ws = wb['Sheet1']
    if 'Sheet1' in wb.sheetnames:
        last_row = ws.max_row
        trackerList = []
        for i in range(1, last_row + 1, 1):
            trackerList.append(ws[f'A{i}'].value)
    return trackerList

def excel_prsnlList_input():

    def get_last_row_in_column(ws, column_letter):
        last_row = ws.max_row
        for row in range(last_row, 0, -1):  # 마지막 행에서부터 거꾸로 탐색
            if ws[f'{column_letter}{row}'].value is not None:
                return row
        return None  # 값이 없다면 None 반환

    prsnlList_path = r"C:\Users\xten\Desktop\prsnlList.xlsx"
    # prsnlList_path = r"C:\Users\kfri1\Desktop\PersonalInfoList.xlsx"

    wb = openpyxl.load_workbook(prsnlList_path)
    ws = wb['Sheet1']

    last_row_in_A = get_last_row_in_column(ws, 'A')
    last_row_in_B = get_last_row_in_column(ws, 'B')


    # if 'Sheet1' in wb.sheetnames:
    #     last_row = ws.max_row
    #     prsnlList = []
    #     for i in range(1, last_row + 1, 1):
    #         prsnlList.append(str(ws[f'A{i}'].value))        

    total_prsnlList = []
    key_prsnlList = []
    value_prsnlList = []

    for i in range(1, last_row_in_A + 1, 1):
        total_prsnlList.append(str(ws[f'A{i}'].value))
        key_prsnlList.append(str(ws[f'A{i}'].value))

    for i in range(1, last_row_in_B + 1, 1):
        total_prsnlList.append(str(ws[f'B{i}'].value))
        value_prsnlList.append(str(ws[f'B{i}'].value))

    # print(len(total_prsnlList))

    return total_prsnlList, key_prsnlList, value_prsnlList




def match_trackerList(trackerList, host):
    for pattern in trackerList:
        # trackerList의 각 항목을 단어로 검색하기 위해 단어 경계(\b)를 추가
        word_pattern = rf'{re.escape(pattern)}\b'
        if re.search(word_pattern, host, re.IGNORECASE):
            return True
    return False



def match_prsnlList(prsnlList, kv, matched_patterns, data_to_write):


    exception_keyword = 'x86'


    # Check for gzip or deflate compressed data and decompress if necessary
    if isinstance(kv, bytes):
        try:
            # Try decompressing with gzip
            with gzip.GzipFile(fileobj=BytesIO(kv)) as f:
                kv = f.read().decode('utf-8')
        except (OSError, gzip.BadGzipFile):
            try:
                # Try decompressing with zlib
                kv = zlib.decompress(kv).decode('utf-8')
            except zlib.error:
                # If neither gzip nor zlib decompression works, assume it's regular bytes
                kv = kv.decode('utf-8', errors='ignore')

    if isinstance(kv, list) and len(kv) == 2:
        k,v = kv

        if isinstance(k, str):
            # try:
            #     v = json.loads(v)
            #     print("JSON 데이터:", v)
            # except json.JSONDecodeError as e:
            #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)
            try:
                # 문자열을 평가하여 리스트인지 확인
                k = ast.literal_eval(k)
                # 평가된 결과가 리스트인지 확인
            except (ValueError, SyntaxError) as e:
                # 평가 과정에서 오류가 발생하면 리스트가 아님
                print(e)

        if isinstance(k, (list, dict, tuple)):
            matched_patterns, data_to_write = match_prsnlList(prsnlList, k, matched_patterns, data_to_write)
            if matched_patterns is None:
                return None, data_to_write

        else:

            if isinstance(k, str) and k.startswith('='):
                data_to_write.append("\\" + k +": ")
            else:
                data_to_write.append(str(k)+": ")

            # data_to_write.append(str(k)+": ")

            for pattern in prsnlList:
                if pattern is None:  # None 무시
                    continue

                word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환

                if re.search(word_pattern, str(k), re.IGNORECASE):
                    matched_patterns.append(pattern)

                    # print(f"world_pattern = {word_pattern}")
                    # print(f"@@@{pattern}@@@")

                if re.search(exception_keyword, str(k), re.IGNORECASE):
                    return None, data_to_write

        if isinstance(v, str):
            # try:
            #     v = json.loads(v)
            #     print("JSON 데이터:", v)
            # except json.JSONDecodeError as e:
            #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)
            try:
                # 문자열을 평가하여 리스트인지 확인
                v = ast.literal_eval(v)
                # 평가된 결과가 리스트인지 확인
            except (ValueError, SyntaxError) as e:
                # 평가 과정에서 오류가 발생하면 리스트가 아님
                print(e)


        if isinstance(v, (list, dict, tuple)):
            matched_patterns, data_to_write = match_prsnlList(prsnlList, v, matched_patterns, data_to_write)
            if matched_patterns is None:
                return None, data_to_write
            
        else:

            if isinstance(v, str) and v.startswith('='):
                data_to_write.append("\\" + v)
            else:
                data_to_write.append(str(v))


            # data_to_write.append(str(v))

            for pattern in prsnlList:
                if pattern is None:  # None 무시
                    continue
                # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'
                word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환

                if re.search(word_pattern, str(v), re.IGNORECASE):
                    matched_patterns.append(pattern)

                    # print(f"world_pattern = {word_pattern}")
                    # print(f"@@@{pattern}@@@")

                if re.search(exception_keyword, str(v), re.IGNORECASE):
                    return None, data_to_write

    elif isinstance(kv, list):
        for item in kv:
            # print(item)
            if isinstance(item, str):
                # try:
                #     item = json.loads(item)
                #     print("JSON 데이터:", item)
                # except json.JSONDecodeError as e:
                #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)

                try:
                    # 문자열을 평가하여 리스트인지 확인
                    item = ast.literal_eval(item)
                    # 평가된 결과가 리스트인지 확인
                except (ValueError, SyntaxError) as e:
                    # 평가 과정에서 오류가 발생하면 리스트가 아님
                    print(e)

            if isinstance(item, (list, dict, tuple)):
                matched_patterns, data_to_write = match_prsnlList(prsnlList, item, matched_patterns, data_to_write)
                if matched_patterns is None:
                    return None, data_to_write

            else:

                if isinstance(item, str) and item.startswith('='):
                    data_to_write.append("\\" + item)
                else:
                    data_to_write.append(str(item))

                # data_to_write.append(str(item))
                for pattern in prsnlList:
                    
                    if pattern is None:  # None 무시
                        continue
                    word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'
                    # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'
 
                    if re.search(word_pattern, str(item), re.IGNORECASE):
                        print(f"world_pattern = {word_pattern}")

                        # matched_patterns.append(pattern)
                        # print(f"@@@{pattern}@@@")

                    if re.search(exception_keyword, str(item), re.IGNORECASE):
                        return None, data_to_write
        
    elif isinstance(kv, tuple) and len(kv) == 2:
        k,v = kv

        if isinstance(k, str):
            # try:
            #     v = json.loads(v)
            #     print("JSON 데이터:", v)
            # except json.JSONDecodeError as e:
            #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)
            try:
                # 문자열을 평가하여 리스트인지 확인
                k = ast.literal_eval(k)
                # 평가된 결과가 리스트인지 확인
            except (ValueError, SyntaxError) as e:
                # 평가 과정에서 오류가 발생하면 리스트가 아님
                print(e)
        
        if isinstance(k, (list, dict, tuple)):
            matched_patterns, data_to_write = match_prsnlList(prsnlList, k, matched_patterns, data_to_write)
            if matched_patterns is None:
                return None, data_to_write     
        else:


            if isinstance(k, str) and k.startswith('='):
                data_to_write.append("\\" + k +": ")
            else:
                data_to_write.append(str(k)+": ")

            # data_to_write.append(str(k)+": ")
       
            for pattern in prsnlList:
                if pattern is None:  # None 무시
                    continue

                word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환

                if re.search(word_pattern, str(k), re.IGNORECASE):
                    matched_patterns.append(pattern)

                    # print(f"world_pattern = {word_pattern}")
                    # print(f"@@@{pattern}@@@")

                if re.search(exception_keyword, str(k), re.IGNORECASE):
                    return None, data_to_write

        if isinstance(v, str):
            # try:
            #     v = json.loads(v)
            #     print("JSON 데이터:", v)
            # except json.JSONDecodeError as e:
            #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)
            try:
                # 문자열을 평가하여 리스트인지 확인
                v = ast.literal_eval(v)
                # 평가된 결과가 리스트인지 확인
            except (ValueError, SyntaxError) as e:
                # 평가 과정에서 오류가 발생하면 리스트가 아님
                print(e)


        if isinstance(v, (list, dict, tuple)):
            matched_patterns, data_to_write = match_prsnlList(prsnlList, v, matched_patterns, data_to_write)
            if matched_patterns is None:
                return None, data_to_write
            
        else:

            if isinstance(v, str) and v.startswith('='):
                data_to_write.append("\\" + v)
            else:
                data_to_write.append(str(v))


            # data_to_write.append(str(v))

            for pattern in prsnlList:
                if pattern is None:  # None 무시
                    continue
                # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'
                word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환

                if re.search(word_pattern, str(v), re.IGNORECASE):
                    matched_patterns.append(pattern)

                    # print(f"world_pattern = {word_pattern}")
                    # print(f"@@@{pattern}@@@")

                if re.search(exception_keyword, str(v), re.IGNORECASE):
                    return None, data_to_write

    elif isinstance(kv, tuple):
        for item in kv:

            if isinstance(item, str):
                # try:
                #     item = json.loads(item)
                #     print("JSON 데이터:", item)
                # except json.JSONDecodeError as e:
                #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)
                try:
                    # 문자열을 평가하여 리스트인지 확인
                    item = ast.literal_eval(item)
                    # 평가된 결과가 리스트인지 확인
                except (ValueError, SyntaxError) as e:
                    # 평가 과정에서 오류가 발생하면 리스트가 아님
                    print(e)


            if isinstance(item, (list, dict, tuple)):
                matched_patterns, data_to_write = match_prsnlList(prsnlList, item, matched_patterns, data_to_write)
                if matched_patterns is None:
                    return None, data_to_write
                
            else:

                if isinstance(item, str) and item.startswith('='):
                    data_to_write.append("\\" + item)
                else:
                    data_to_write.append(str(item))
                    
                # data_to_write.append(str(item))
                for pattern in prsnlList:
                    if pattern is None:  # None 무시
                        continue
                    word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환
                    # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'

                    if re.search(word_pattern, str(item), re.IGNORECASE):
                        matched_patterns.append(pattern)

                        # print(f"world_pattern = {word_pattern}")
                        # print(f"@@@{pattern}@@@")

                    if re.search(exception_keyword, str(item), re.IGNORECASE):
                        return None, data_to_write
                
    elif isinstance(kv, dict):
        for k, v in kv.items():


            if isinstance(k, str) and k.startswith('='):
                data_to_write.append("\\" + k +": ")
            else:
                data_to_write.append(str(k)+": ")
            
            # data_to_write.append(str(k)+": ")


            for pattern in prsnlList:
                if pattern is None:  # None 무시
                    continue

                word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환

                if re.search(word_pattern, str(k), re.IGNORECASE):
                    matched_patterns.append(pattern)

                    # print(f"world_pattern = {word_pattern}")
                    # print(f"@@@{pattern}@@@")

                if re.search(exception_keyword, str(k), re.IGNORECASE):
                    return None, data_to_write

            if isinstance(v, str):
                # try:
                #     v = json.loads(v)
                #     print("JSON 데이터:", v)
                # except json.JSONDecodeError as e:
                #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)            

                try:
                    # 문자열을 평가하여 리스트인지 확인
                    v = ast.literal_eval(v)
                    # 평가된 결과가 리스트인지 확인
                except (ValueError, SyntaxError) as e:
                    # 평가 과정에서 오류가 발생하면 리스트가 아님
                    print(e)

            if isinstance(v, (list, dict, tuple)):

                matched_patterns, data_to_write = match_prsnlList(prsnlList, v, matched_patterns, data_to_write)
                if matched_patterns is None:
                    return None, data_to_write
                    
            else:

                if isinstance(v, str) and v.startswith('='):
                    data_to_write.append("\\" + v)
                else:
                    data_to_write.append(str(v))

                # data_to_write.append(str(v))

                for pattern in prsnlList:
                    if pattern is None:  # None 무시
                        continue
                    
                    word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'  # 패턴을 문자열로 변환

                    # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'

                    if re.search(word_pattern, str(v), re.IGNORECASE):
                        matched_patterns.append(pattern)

                        # print(f"world_pattern = {word_pattern}")
                        # print(f"@@@{pattern}@@@")


                    if re.search(exception_keyword, str(v), re.IGNORECASE):
                        return None, data_to_write
                
    elif isinstance(kv, str):

        if isinstance(kv, str):
            # try:
            #     kv = json.loads(kv)
            #     print("JSON 데이터:", kv)
            # except json.JSONDecodeError as e:
            #     print("JSON형태가 아니거나 JSON 파싱 오류:", e)
            try:
                # 문자열을 평가하여 리스트인지 확인
                kv = ast.literal_eval(kv)
                # 평가된 결과가 리스트인지 확인
            except (ValueError, SyntaxError) as e:
                # 평가 과정에서 오류가 발생하면 리스트가 아님
                print(e)            

        if isinstance(kv, (list, dict, tuple)):

            matched_patterns, data_to_write = match_prsnlList(prsnlList, kv, matched_patterns, data_to_write)
            if matched_patterns is None:
                return None, data_to_write

        else:


            if isinstance(kv, str) and kv.startswith('='):
                data_to_write.append("\\" + kv)
            else:
                data_to_write.append(str(kv))

            # data_to_write.append(str(kv))

            for pattern in prsnlList:
                if pattern is None:  # None 무시
                    continue
                word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)' 
                # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'

                if re.search(word_pattern, str(kv), re.IGNORECASE):
                    matched_patterns.append(pattern)

                    # print(f"world_pattern = {word_pattern}")
                    # print(f"@@@{pattern}@@@")

                if re.search(exception_keyword, str(kv), re.IGNORECASE):
                    return None, data_to_write

    return matched_patterns, data_to_write



def clean_host_name(host):
    """
    호스트 이름을 Excel 시트 이름으로 사용할 수 있도록 수정합니다.
    길이를 31자로 제한하고 특수 문자를 제거합니다.
    """
    # 특수 문자 제거
    clean_host = re.sub(r'[\/:*?"<>|]', '', host)
    
    # 길이 제한
    if len(clean_host) > MAX_HOSTNAME_LENGTH:
        clean_host = clean_host[:MAX_HOSTNAME_LENGTH]
    
    return clean_host


def clean_string(value):
    # 허용되지 않는 제어 문자를 제거

    # 먼저 value가 문자열인지 확인합니다.
    if not isinstance(value, (str, bytes)):
        # 문자열이나 바이트가 아니라면, 문자열로 변환 시도
        value = str(value)
    return re.sub(r'[\x00-\x1F\x7F]', '', value)


def open_excel(app_name, key_prsnlList):
    
    # results_folder_path = r"C:\Users\kfri1\Desktop\testing2"
    results_folder_path = r"C:\Users\xten\Desktop\testing4"


    result_path = os.path.join(results_folder_path, f"{app_name}.xlsx")

    if os.path.exists(result_path):
        wb = openpyxl.load_workbook(result_path)
    else:
        wb = openpyxl.Workbook()

        # # 기본으로 생성되는 'Sheet' 시트를 제거
        # default_sheet = wb.active
        # wb.remove(default_sheet)

        # 기본으로 생성되는 'Sheet' 시트의 이름을 'result'로 변경
        wb.active.title = 'Result'
        ws = wb['Result']

    for col, key in enumerate(key_prsnlList, start=5):  # E열은 5번째 열이므로 start=5
        key_prsnlList_cell = ws.cell(row=1, column=col, value=key)
        key_prsnlList_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    ws.cell(row = 1, column = 4 + len(key_prsnlList) + 1, value = 'ETC')
    
    return wb, result_path

def write_to_excel(host, data, total_prsnlList, no_dup_matched_patterns_to_write, wb):


    clean_host = clean_host_name(host)

    if clean_host in wb.sheetnames:
        ws = wb[clean_host]
    else:
        ws = wb.create_sheet(title=clean_host)
    # 현재 마지막 열을 찾기
    last_column = ws.max_column
    if last_column == 1 and ws.cell(row=1, column=1).value is None:
        last_column = 0

    # 새로운 열에서 데이터 입력 시작
    start_column = last_column + 1
    for i, value in enumerate(data, start=1):
        try:
            cleaned_value = clean_string(value)
            cell = ws.cell(row=i, column=start_column, value=cleaned_value)

            print(value)

            if total_prsnlList:
                for pattern in total_prsnlList:

                    word_pattern = rf'(?<!\w){re.escape(str(pattern))}(?!\w)'
                    # word_pattern = rf'(?!\w|_|-){re.escape(str(pattern))}(?!\w|_|-)'
                    if re.search(word_pattern, value, re.IGNORECASE):
                        cell.font = openpyxl.styles.Font(bold=True, color="FF0000")  # 개별 셀에 폰트 스타일 적용            




        except IllegalCharacterError as e:
            print(f"Illegal character in value: {value}. Error: {e}")

 
    if no_dup_matched_patterns_to_write:

        last_row_in_column = ws.max_row

        while last_row_in_column > 0 and ws.cell(row=last_row_in_column, column=start_column).value is None:
            last_row_in_column -= 1

        next_row = last_row_in_column + 1
        for j, pattern in enumerate(no_dup_matched_patterns_to_write, start=0):
            pattern_cell = ws.cell(row=next_row + j, column=start_column, value=pattern)
            pattern_cell.font = openpyxl.styles.Font(color="0000FF")  # 파란색 텍스트로 설정




def write_Result(host, hostlist, wb, key_prsnlList, value_prsnlList, no_dup_matched_patterns, row_hostlist_number, hostlist_number_dict, hostlist_etc_dict, hostlist_count, app_name):

    ws = wb['Result']
    if host not in hostlist:
        hostlist.append(host)
        hostlist_number_dict[host] = row_hostlist_number
        hostlist_cell = ws.cell(row = row_hostlist_number, column = 4, value = host)
        hostlist_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        # if host not in hostlist_etc_dict:
        #     hostlist_etc_dict[host] = set()
        hostlist_etc_dict[host] = set()
        hostlist_count_cell = ws.cell(row = row_hostlist_number, column = 1, value = hostlist_count)
        hostlist_count_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        app_name_cell = ws.cell(row = row_hostlist_number, column = 2, value = app_name)
        app_name_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        row_hostlist_number += 1
        hostlist_count +=1
    for col in range(5, 5 + len(key_prsnlList)):
        for matched_pattern in no_dup_matched_patterns:
            if ws.cell(row = 1, column = col).value == matched_pattern:
                #값이 있을 경우 덮어씌워버림
                key_checked_cell = ws.cell(row = hostlist_number_dict[host], column = col, value = 'O')
                key_checked_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')



    for matched_pattern in no_dup_matched_patterns:
        for value in value_prsnlList:
            if matched_pattern == value:
                # etc_value.add(matched_pattern)
                hostlist_etc_dict[host].add(matched_pattern)

    
    # hostlist_etc_dict[host] = 
    # etc_value_as_string = ', '.join(etc_value)
    hostlist_etc_dict_as_string = ', '.join(hostlist_etc_dict[host])
    ws.cell(row = hostlist_number_dict[host], column = 4 + len(key_prsnlList) + 1, value = hostlist_etc_dict_as_string)


    return hostlist, row_hostlist_number, hostlist_number_dict, hostlist_etc_dict, hostlist_count
    


#===========================================================================================================================================================================================================#

def read_excel_data():
    excel_data_path = input("excel_data 경로를 입력해주세요: ")
    wb = openpyxl.load_workbook(excel_data_path)
    ws = wb['Sheet1']
    ranking = ws[f'A{row}'].value if ws[f'A{row}'].value is not None else ''


    
